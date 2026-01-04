# excel_template_generator.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

class ExcelTemplateGenerator:
    def __init__(self, output_path="dashboard_config_template.xlsx"):
        self.output_path = output_path
        self.wb = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Configure les styles Excel"""
        self.colors = {
            'primary': 'FF007348',
            'secondary': 'FF00A678',
            'accent': 'FF8DC9AB',
            'light': 'FFF9FDFB',
            'header': 'FFE8F4F0'
        }
        
        self.header_fill = PatternFill(
            start_color=self.colors['header'],
            end_color=self.colors['header'],
            fill_type='solid'
        )
        
        self.header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.colors['primary']
        )
        
        self.header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        self.required_fill = PatternFill(
            start_color='FFFFE0E0',
            end_color='FFFFE0E0',
            fill_type='solid'
        )
        
        self.example_fill = PatternFill(
            start_color='FFF0F9F5',
            end_color='FFF0F9F5',
            fill_type='solid'
        )
        
        self.example_font = Font(
            name='Calibri',
            size=10,
            italic=True,
            color='FF666666'
        )
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_sheet_pages(self):
        """Crée la feuille Pages"""
        ws = self.wb.create_sheet("Pages")
        
        headers = [
            "page_id", "page_name", "description", "layout_template", 
            "theme", "parent_page", "order", "enabled", "refresh_interval"
        ]
        
        descriptions = [
            "ID unique de la page (obligatoire)",
            "Nom affiché (obligatoire)",
            "Description de la page",
            "Template de layout",
            "Thème de couleur",
            "Page parent pour navigation",
            "Ordre d'affichage",
            "Activé (TRUE/FALSE)",
            "Intervalle rafraîchissement (secondes)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["EQ_DASH_1", "Dashboard Actions", "Analyse des actions", "standard", "green", "", 1, True, 300],
            ["CREDIT_A", "Dashboard Crédit A", "Analyse crédit", "two_columns", "blue", "", 2, True, 600],
            ["STRESS_1", "Test de Stress", "Scénarios stress", "full", "dark", "", 3, True, 900],
            ["CORR_1", "Corrélation", "Matrices corrélation", "grid", "custom", "", 4, True, 1200]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = self.example_font
                cell.border = self.thin_border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        print("✅ Feuille 'Pages' créée")
    
    def create_sheet_sections(self):
        """Crée la feuille Sections"""
        ws = self.wb.create_sheet("Sections")
        
        headers = [
            "page_id", "section_id", "section_type", "title", "description", 
            "layout", "columns", "order", "collapsible", "default_collapsed"
        ]
        
        descriptions = [
            "ID de la page parent",
            "ID unique de la section",
            "Type (container, section, subsection)",
            "Titre affiché",
            "Description",
            "Layout (vertical, horizontal, grid)",
            "Nombre de colonnes",
            "Ordre dans la page",
            "Rétractable (TRUE/FALSE)",
            "Réduit par défaut (TRUE/FALSE)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["EQ_DASH_1", "filters", "container", "Filtres", "Filtres principaux", "horizontal", 3, 1, True, False],
            ["EQ_DASH_1", "kpis", "container", "Indicateurs", "KPI principaux", "grid", 4, 2, True, False],
            ["EQ_DASH_1", "charts", "section", "Graphiques", "Visualisations", "columns", 2, 3, True, False],
            ["EQ_DASH_1", "table", "container", "Tableau", "Données détaillées", "full", 1, 4, True, True]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = self.example_font
                cell.border = self.thin_border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        print("✅ Feuille 'Sections' créée")
    
    def create_sheet_components(self):
        """Crée la feuille Components"""
        ws = self.wb.create_sheet("Components")
        
        headers = [
            "page_id", "section_id", "component_id", "component_type", "title",
            "data_source_type", "data_source_query", "properties_json",
            "order", "width", "height"
        ]
        
        descriptions = [
            "ID de la page",
            "ID de la section",
            "ID unique du composant",
            "Type (kpi, chart, table, filter, form, text)",
            "Titre du composant",
            "Type source (sql, static, api)",
            "Requête SQL ou données",
            "Propriétés JSON",
            "Ordre dans la section",
            "Largeur (1-12 ou pixels)",
            "Hauteur en pixels"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["EQ_DASH_1", "kpis", "kpi_total", "kpi", "Valeur Totale", "sql", 
             "SELECT SUM(value) FROM equity", '{"format": "currency", "size": "large"}', 1, 3, 150],
            
            ["EQ_DASH_1", "charts", "chart_trend", "chart", "Évolution", "sql", 
             "SELECT date, value FROM equity_daily", '{"type": "line", "height": 400}', 1, 6, 400],
            
            ["EQ_DASH_1", "table", "table_data", "table", "Positions", "sql", 
             "SELECT * FROM equity_positions", '{"pagination": true, "page_size": 10}', 1, 12, 500],
            
            ["EQ_DASH_1", "filters", "filter_region", "filter", "Filtre Région", "static", 
             "", '{"type": "dropdown", "options": ["US", "EU", "ASIA"]}', 1, 4, "auto"]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = self.example_font
                cell.border = self.thin_border
        
        col_widths = [12, 15, 20, 12, 25, 12, 50, 40, 8, 8, 8]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        print("✅ Feuille 'Components' créée")
    
    def create_sheet_data_sources(self):
        """Crée la feuille Data_Sources"""
        ws = self.wb.create_sheet("Data_Sources")
        
        headers = [
            "component_id", "data_type", "query", "cache_ttl", "refresh_interval"
        ]
        
        descriptions = [
            "ID du composant lié",
            "Type (sql, static, api)",
            "Requête SQL ou données",
            "Cache en secondes",
            "Rafraîchissement auto (secondes)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["kpi_total", "sql", "SELECT SUM(value) as total FROM equity_holdings WHERE date = CURRENT_DATE", 300, 60],
            ["chart_trend", "sql", "SELECT date, sector, SUM(value) as total FROM equity_daily WHERE date >= CURRENT_DATE - 30 GROUP BY date, sector", 600, 300],
            ["table_data", "sql", "SELECT * FROM v_equity_positions ORDER BY value DESC", 900, 600]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = Font(size=9)
                cell.border = self.thin_border
                cell.alignment = Alignment(wrap_text=True)
        
        col_widths = [20, 12, 60, 10, 10]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        print("✅ Feuille 'Data_Sources' créée")
    
    def create_sheet_filters(self):
        """Crée la feuille Filters"""
        ws = self.wb.create_sheet("Filters")
        
        headers = [
            "page_id", "filter_id", "filter_type", "label", "data_source_value",
            "default_value", "multiple", "target_components"
        ]
        
        descriptions = [
            "ID de la page",
            "ID unique du filtre",
            "Type (dropdown, checklist, date_range)",
            "Libellé affiché",
            "Options (JSON ou SQL)",
            "Valeur par défaut",
            "Choix multiple (TRUE/FALSE)",
            "Composants affectés (IDs séparés par virgules)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["EQ_DASH_1", "filter_region", "dropdown", "Région", '["US", "EU", "ASIA"]', 
             '["US", "EU"]', True, "kpi_total,chart_trend,table_data"],
            
            ["EQ_DASH_1", "filter_time", "date_range", "Période", '["7d", "30d", "90d"]', 
             "30d", False, "chart_trend,table_data"],
            
            ["EQ_DASH_1", "filter_sector", "dropdown", "Secteur", "SELECT DISTINCT sector FROM sectors", 
             "", True, "table_data"]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = self.example_font
                cell.border = self.thin_border
        
        col_widths = [15, 15, 12, 20, 40, 20, 8, 30]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        print("✅ Feuille 'Filters' créée")
    
    def create_sheet_actions(self):
        """Crée la feuille Actions"""
        ws = self.wb.create_sheet("Actions")
        
        headers = [
            "component_id", "action_type", "action_label", "endpoint",
            "target_components", "enabled"
        ]
        
        descriptions = [
            "ID du composant",
            "Type (click, drilldown, export, refresh)",
            "Libellé du bouton",
            "URL ou fonction",
            "Composants à rafraîchir",
            "Activé (TRUE/FALSE)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            ws.cell(row=2, column=col, value=descriptions[col-1])
        
        examples = [
            ["kpi_total", "drilldown", "Voir détails", "/api/equity/details", "", True],
            ["table_data", "export", "Exporter Excel", "/api/export/excel", "", True],
            ["filter_region", "change", "", "", "kpi_total,chart_trend,table_data", True]
        ]
        
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = self.example_fill
                cell.font = self.example_font
                cell.border = self.thin_border
        
        col_widths = [20, 12, 15, 25, 25, 8]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        print("✅ Feuille 'Actions' créée")
    
    def create_sheet_documentation(self):
        """Crée la feuille Documentation"""
        ws = self.wb.create_sheet("Documentation")
        
        # Titre
        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value="📋 GUIDE DE CONFIGURATION")
        title_cell.font = Font(size=16, bold=True, color=self.colors['primary'])
        title_cell.alignment = Alignment(horizontal='center')
        
        # Sections
        sections = [
            ("📌 INSTRUCTIONS", [
                "1. Remplissez les feuilles dans l'ordre",
                "2. Les champs en rouge sont obligatoires",
                "3. Suivez les exemples fournis",
                "4. Conservez les IDs uniques"
            ]),
            
            ("🔗 ORDRE", [
                "1. Pages → 2. Sections → 3. Components",
                "4. Data_Sources → 5. Filters → 6. Actions"
            ]),
            
            ("🎨 COULEURS", [
                "Primaire: #007348",
                "Secondaire: #00A678", 
                "Accent: #8DC9AB"
            ]),
            
            ("📊 COMPOSANTS", [
                "kpi: Indicateurs",
                "chart: Graphiques",
                "table: Tableaux",
                "filter: Filtres"
            ])
        ]
        
        row = 3
        for section_title, items in sections:
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws.cell(row=row, column=1, value=section_title)
            section_cell.font = Font(size=13, bold=True, color=self.colors['primary'])
            row += 1
            
            for item in items:
                ws.cell(row=row, column=2, value="• " + item)
                row += 1
            
            row += 1
        
        # Largeurs
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 5
        
        print("✅ Feuille 'Documentation' créée")
    
    def generate_template(self):
        """Génère le template complet"""
        print("🔧 Génération du template Excel...")
        
        # Supprimer feuille par défaut
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Créer les feuilles
        self.create_sheet_documentation()
        self.create_sheet_pages()
        self.create_sheet_sections()
        self.create_sheet_components()
        self.create_sheet_data_sources()
        self.create_sheet_filters()
        self.create_sheet_actions()
        
        # Sauvegarder
        self.wb.save(self.output_path)
        
        print(f"✅ Template généré: {self.output_path}")
        print(f"📊 Feuilles: {len(self.wb.sheetnames)}")
        
        return self.output_path

# ====== CONVERTISSEUR EXCEL → JSON ======

class ExcelToJsonConverter:
    """Convertit Excel en JSON"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.data = {}
    
    def load_excel(self):
        """Charge le fichier Excel"""
        print(f"📖 Chargement: {self.excel_path}")
        
        try:
            xls = pd.ExcelFile(self.excel_path)
            
            for sheet_name in xls.sheet_names:
                if sheet_name != 'Documentation':
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    df = df.dropna(how='all')
                    self.data[sheet_name] = df
                    print(f"  ✅ {sheet_name}: {len(df)} lignes")
            
            return True
            
        except Exception as e:
            print(f"❌ Erreur: {e}")
            return False
    
    def convert_value(self, value):
        """Convertit les valeurs"""
        if pd.isna(value):
            return None
        
        if isinstance(value, str):
            value = value.strip()
            
            if value.upper() == 'TRUE':
                return True
            elif value.upper() == 'FALSE':
                return False
            
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except:
                pass
            
            if value.startswith('[') and value.endswith(']'):
                try:
                    return json.loads(value)
                except:
                    pass
            
            if value.startswith('{') and value.endswith('}'):
                try:
                    return json.loads(value)
                except:
                    pass
        
        return value
    
    def generate_json(self, output_dir="page_configs"):
        """Génère les fichiers JSON"""
        print(f"🔄 Génération JSON dans: {output_dir}")
        
        import os
        os.makedirs(output_dir, exist_ok=True)
        
        if 'Pages' not in self.data:
            print("❌ Aucune page trouvée")
            return []
        
        pages_df = self.data['Pages']
        generated_files = []
        
        for _, page in pages_df.iterrows():
            if page.get('enabled', True) != True:
                continue
            
            page_id = page['page_id']
            page_config = self.build_page_config(page_id, page)
            
            filename = f"{page_id}.json"
            filepath = os.path.join(output_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(page_config, f, indent=2, ensure_ascii=False)
            
            print(f"  ✅ {filename}")
            generated_files.append(filepath)
        
        print(f"🎉 {len(generated_files)} fichiers générés")
        return generated_files
    
    def build_page_config(self, page_id, page_data):
        """Construit la config d'une page"""
        config = {
            "type": "dashboard",
            "id": page_id,
            "title": page_data.get('page_name', page_id),
            "description": page_data.get('description', ''),
            "layout": page_data.get('layout_template', 'vertical'),
            "refresh_interval": page_data.get('refresh_interval', 300),
            "sections": [],
            "metadata": {
                "created_at": datetime.now().isoformat(),
                "source": self.excel_path
            }
        }
        
        if 'theme' in page_data and pd.notna(page_data['theme']):
            config['theme'] = page_data['theme']
        
        if 'Sections' in self.data:
            sections_df = self.data['Sections']
            page_sections = sections_df[sections_df['page_id'] == page_id]
            
            for _, section in page_sections.iterrows():
                section_config = self.build_section_config(section)
                config['sections'].append(section_config)
        
        return config
    
    def build_section_config(self, section_data):
        """Construit la config d'une section"""
        section_config = {
            "id": section_data['section_id'],
            "type": section_data.get('section_type', 'container'),
            "title": section_data.get('title', ''),
            "description": section_data.get('description', ''),
            "collapsible": section_data.get('collapsible', True),
            "default_collapsed": section_data.get('default_collapsed', False),
            "layout": section_data.get('layout', 'vertical'),
            "components": []
        }
        
        if pd.notna(section_data.get('columns')):
            section_config['columns'] = int(section_data['columns'])
        
        if 'Components' in self.data:
            components_df = self.data['Components']
            section_components = components_df[
                (components_df['page_id'] == section_data['page_id']) & 
                (components_df['section_id'] == section_data['section_id'])
            ]
            
            for _, component in section_components.iterrows():
                component_config = self.build_component_config(component)
                section_config['components'].append(component_config)
        
        return section_config
    
    def build_component_config(self, component_data):
        """Construit la config d'un composant"""
        config = {
            "id": component_data['component_id'],
            "type": component_data.get('component_type', 'text'),
            "title": component_data.get('title', ''),
            "order": int(component_data.get('order', 1))
        }
        
        properties = component_data.get('properties_json')
        if pd.notna(properties):
            if isinstance(properties, str):
                try:
                    config['config'] = json.loads(properties)
                except:
                    config['config'] = properties
            else:
                config['config'] = properties
        
        if pd.notna(component_data.get('data_source_type')):
            data_config = {
                "type": component_data['data_source_type']
            }
            
            if pd.notna(component_data.get('data_source_query')):
                data_config['query'] = component_data['data_source_query']
            
            config['data'] = data_config
        
        if pd.notna(component_data.get('width')):
            config['width'] = component_data['width']
        
        if pd.notna(component_data.get('height')):
            config['height'] = component_data['height']
        
        return config

# ====== MAIN ======

def main():
    """Fonction principale"""
    print("=" * 50)
    print("📊 EXCEL TEMPLATE GENERATOR")
    print("=" * 50)
    print("\n1. 🔧 Générer template Excel")
    print("2. 🔄 Convertir Excel en JSON")
    print("0. ❌ Quitter")
    
    choice = input("\n👉 Choix: ").strip()
    
    if choice == '1':
        generator = ExcelTemplateGenerator("dashboard_config.xlsx")
        generator.generate_template()
        print("\n✅ Template prêt à être rempli!")
        
    elif choice == '2':
        excel_file = input("📂 Fichier Excel: ").strip()
        if not excel_file:
            excel_file = "dashboard_config.xlsx"
        
        converter = ExcelToJsonConverter(excel_file)
        if converter.load_excel():
            converter.generate_json()
        
    elif choice == '0':
        print("👋 Au revoir!")
        return
    
    else:
        print("❌ Choix invalide")

if __name__ == "__main__":
    try:
        import openpyxl
        import pandas as pd
    except ImportError:
        print("📦 Installation des dépendances...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pandas"])
        print("✅ Relancez le script")
        sys.exit(1)
    
    main()